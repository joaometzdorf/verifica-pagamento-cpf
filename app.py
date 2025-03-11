import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# Entrar na planilha e extrair o cpf do cliente
dados_clientes = openpyxl.load_workbook("dados_clientes.xlsx")
planilha_clientes = dados_clientes["Sheet1"]

# Entro no site https://consultcpf-devaprender.netlify.app/
driver = webdriver.Chrome()
driver.maximize_window()
driver.get("https://consultcpf-devaprender.netlify.app/")
sleep(2)

# Pego o cpf da planilha para pesquisar o status de pagamento daquele cliente
for row in planilha_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = row
    input_cpf = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    input_cpf.send_keys(cpf)

    button_submit = driver.find_element(
        By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']"
    )
    button_submit.click()
    sleep(4)
    input_cpf.clear()
    sleep(1)

    # Verificar se está "em dia" ou "atrasado"
    status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
    # Se estiver "em dia", pegar data do pagamento e método de pagamento
    if status.text == "em dia":
        fechamento = openpyxl.load_workbook("planilha_fechamento.xlsx")
        planilha_fechamento = fechamento["Sheet1"]
        data_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")
        # Inserir essas novas informações(nome, valor, cpf, vencimento, status e caso esteja em dia, data pagamento, método pagamento(cartão ou boleto))
        planilha_fechamento.append(
            [
                nome,
                valor,
                cpf,
                vencimento,
                status.text,
                data_pagamento.text.split()[3],
                metodo_pagamento.text.split()[3],
            ]
        )
        fechamento.save("planilha_fechamento.xlsx")
    # Caso contrário(se estiver atrasado), colocar o status como pendente
    else:
        fechamento = openpyxl.load_workbook("planilha_fechamento.xlsx")
        planilha_fechamento = fechamento["Sheet1"]
        planilha_fechamento.append([nome, valor, cpf, vencimento, "pendente", "", ""])
        fechamento.save("planilha_fechamento.xlsx")
# Repetir até chegar no último cliente
